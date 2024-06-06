import  openpyxl
from  openpyxl.utils import get_column_letter 
from pathlib import Path
from ConfigManager.config_manager import ConfigManager
from work_book import Formula, Workbook

def print_error(string_error):
    print(f"\n ***********************************************************\n \
{string_error}\n \
************************************************************\n")


def main():
    config_file_name = "project_config.toml"
    config_manager = ConfigManager(config_file_name)

    if config_manager.error != 0:
        print_error(f"error in config file {config_file_name}. stoping process")
    
    # config  general
    config_general = config_manager.general_config

    root_data = config_general.root_dir

    # config points id 

    ids_config = config_manager.point_id_config

    id_file_path = root_data / ids_config.id_path

    id_wb_instance = Workbook()

    id_wb_instance.load_workbook(id_file_path)

    if id_wb_instance.error != 0:
        print_error("fatal error stopping process")
        exit()

    id_data_sheet =  id_wb_instance.add_id_sheet(ids_config.sheet_name, ids_config.id_pattern)

    ids_data = id_data_sheet.get_data_by_columns_name(list([ids_config.id_column_title]))

    if ids_data is None:
        print_error(f"something went wrong processing point ids file {id_file_path}. exit process")
        exit()

    id_wb_instance.close_workbook()

    ids = { item : None for item in ids_data[ids_config.id_column_title] }
    
    # config  source data

    source_data_config = config_manager.source_data_config

    source_data_path = root_data / source_data_config.data_path

    # config  target data

    target_data_config = config_manager.target_data_config
    
    target_folder = root_data / target_data_config.folder_path

    #process data sheet

    wb_data = Workbook()

    wb_data.load_workbook(source_data_path)

    if wb_data.error != 0:
        print_error("fatal error stopping process")
        exit()

    target_wb = openpyxl.Workbook()

    for sheet_name in source_data_config.sheets_names:

        columns_names = source_data_config.sheet_to_columns[sheet_name]
        columns_names_target = dict.fromkeys(columns_names, "")

        print(f"\nsheet name: {sheet_name}")

        formula_instances = {}

        #handle columns that have content of formula 
        if sheet_name in target_data_config.sheets.keys():
            for columnName, formula in target_data_config.sheets[sheet_name].items():
                formula_instance = Formula(formula)
                formula_instances[columnName] = formula_instance

                for key in formula_instance.column_dict.keys():
                    if not key in  columns_names_target.keys():
                        columns_names_target.setdefault(key)
                        columns_names.append(key)
                print(f"columns for input {columns_names_target.keys()}")

        data_sheet = wb_data.add_data_sheet(sheet_name, ids, source_data_config.id_column_title)

        data = data_sheet.get_data_by_columns_name(columns_names)

        if data is None:
            print_error("something went wrong in sheet {sheet_name}")
            continue

        target_wb.create_sheet(sheet_name)
        sheet = target_wb[sheet_name]

        sheet_column_to_letters = { title_column : get_column_letter(i + 1) for i, title_column in enumerate(data.keys())}

        num_of_rows = len(data[source_data_config.id_column_title])
        print(f"num of rows {num_of_rows}")

        # replace fromulas column names with excel column letter
        for column_name, formula_instance in formula_instances.items():

            formula_instance.get_processed_formula(sheet_column_to_letters)
            # add destination columns if not exists in data
            if column_name not in data.keys():
                print(f"title {column_name} formula {formula_instance.raw_formula}")
                data[column_name] = range(num_of_rows)

        # write data
        for column, (title_column, values) in enumerate(data.items(), 1):
            sheet.cell(1 , column, title_column) # first row for titles
            for row, val in enumerate(values, 2):
                if title_column in formula_instances.keys():
                    formula_instance = formula_instances[title_column] 

                    formula_with_row = formula_instance.get_formula_with_row_number(row)

                    sheet.cell(row=row, column=column, value=formula_with_row)
                else:
                    sheet.cell(row=row, column=column, value=val)
        

    wb_data.close_workbook()

    target_wb.remove(target_wb.active) #remove auto created sheet

    target_folder.mkdir(parents=True, exist_ok=True)

    target_file = Path(target_data_config.file_prefix + "_" + ids_config.id_column_title +  source_data_path.suffix) if target_data_config.file_prefix else \
                    Path(ids_config.id_column_title +  source_data_path.suffix)
    
    target_full_path = target_folder / target_file
    try:
        target_wb.save(target_full_path)
    except Exception as e:
        print_error(f"{e} Error Saving file {target_full_path}")
        
        exit()

    # success
    print_error(f"Processed Succesfully {config_general.success_strig}")

main()