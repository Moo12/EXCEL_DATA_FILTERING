import re
import  openpyxl
from abc import ABC, abstractmethod
from  openpyxl import styles


class Workbook():
    def __init__(self):
        self.wb_obj = None
        self.error = 0

        self.sheets = {}

    def load_workbook(self, file_path):
        try:
            self.file_path = file_path
            self.wb_obj = openpyxl.load_workbook(file_path)

        except Exception as e:
            print(e)
            self.print_error(file_path)
            self.error = 1

    def close_workbook(self):
        try:
            if self.wb_obj:
                self.wb_obj.close()
        except Exception as e:
            self.print_error(e)
            self.error = 1

    def add_id_sheet(self, sheet_name, pattern):
        if  self.wb_obj:
            sheet_obj = self.wb_obj[sheet_name]

            self.sheets[sheet_name] = SheetId(sheet_obj, sheet_name, pattern)

            return self.sheets[sheet_name]
        else:
            return None
    
    def add_data_sheet(self, sheet_name, ids, id_column_title):
        if  self.wb_obj:
            sheet_obj = self.wb_obj[sheet_name]

            self.sheets[sheet_name] = DataSheet(sheet_obj, sheet_name, ids, id_column_title)

            return self.sheets[sheet_name]
        else:
            return None

    def print_error(error_string):
        print(f" -----------------------------------------------------------\n \
Error in {error_string}. \n \
-----------------------------------------------------------")

class SheetAbstract(ABC):
    def __init__(self, sheet_obj, sheet_name, verbose = False):
        self.sheet_obj = sheet_obj
        self.sheet_name = sheet_name
        self.verbose = verbose

    def is_cell_condition_valid(self, cell):
        return True
    
    def is_row_condition_valid(self, row, columns_index):

        return True

    def get_column_names_to_index(self, column_names):
        column_index_map = dict.fromkeys(column_names, -1)

        flag_reached_title_row = False #flag to skip all blanket

        for row in self.sheet_obj.iter_rows(min_row=1, min_col=0, max_row=self.sheet_obj.max_row, max_col=self.sheet_obj.max_column):
            if not flag_reached_title_row:
                for column in range(self.sheet_obj.max_column):
                    if row[column].value is not None:
                        cell_value = "".join(row[column].value.rstrip().lstrip())
                        if cell_value in column_index_map.keys():
                            flag_reached_title_row = True
                            column_index_map[cell_value] = column
                            print("column title: ",cell_value, "column index: ", column)
            else:
                break

        return column_index_map
    
    def are_all_columns_in_sheet(self, column_index_map, verbose = True):
        columns_not_found = [ column for column, value in column_index_map.items() if value == -1]
        if len(columns_not_found):
            if (verbose):
                Workbook.print_error(f"{columns_not_found} columns were not found in sheet warning")
            return False
        else:
            return True
        
    def get_data_by_columns_name(self, columns_title):
        columns_index = self.get_column_names_to_index(columns_title)

        if not self.sheet_obj:
            Workbook.print_error(f"sheet instance {self.sheet_name} was not created")
            
            return None

        if not self.are_all_columns_in_sheet(columns_index):
            Workbook.print_error(f"could not proocess sheet {self.sheet_name}")

            return None
    
        data = {}

        for row in self.sheet_obj.iter_rows(min_row=2, min_col=0, max_row=self.sheet_obj.max_row, max_col=self.sheet_obj.max_column):
            if self.is_row_condition_valid(row, columns_index):
                for column_name, column_index in columns_index.items():
                    if not column_name in data.keys():
                        data[column_name]= []

                    #print(f"colour: {row[column_index].fill.start_color.index}")
                    if row[column_index].value and self.is_cell_condition_valid(row[column_index]):
                        data[column_name].append(row[column_index].value)

        return data

class SheetId(SheetAbstract):
    def __init__(self, wb_obj, sheet_name, regex_pattern):
        super().__init__(wb_obj, sheet_name)

        self.regex_pattern = regex_pattern

    def is_cell_condition_valid(self, cell):
        cell.value = "".join(cell.value.rstrip().lstrip())
        pattern = re.compile(self.regex_pattern)

        is_cell_fill_white = cell.fill.start_color.index == styles.colors.COLOR_INDEX[0] or cell.fill.end_color.index == styles.colors.COLOR_INDEX[0] 

        if self.verbose:
            print(f"start index {cell.fill.start_color.index} end index {cell.fill.end_color.index}")

        return pattern.match(cell.value) and not is_cell_fill_white

class DataSheet(SheetAbstract):
    def __init__(self, wb_obj, sheet_name, ids, ids_column_name):
        super().__init__(wb_obj, sheet_name)

        self.ids = ids
        self.ids_column_name = ids_column_name

    def is_row_condition_valid(self, row, columns_index):

        if not self.ids_column_name in columns_index.keys():
            raise Exception(f"ids column {self.ids_column} was not found in sheet {self.sheet_name}")
        
        id_column_index = columns_index[self.ids_column_name]

        return row[id_column_index].value and "".join(row[id_column_index].value.rstrip().lstrip()) in self.ids.keys()

class Formula():
    special_words = { "IF": ""}

    def __init__(self, raw_formula, verbose = False):
        self.raw_formula = raw_formula
        formula_columns = re.findall(r'[A-Za-z]+[-_0-9]?[a-zA-Z]', raw_formula) #get all column names from formula string
        self.column_dict = { x : "" for x in formula_columns if x not in Formula.special_words.keys()}
        self.verbose = verbose
    
    def get_processed_formula(self, sheet_column_to_letters):
        for column in self.column_dict.keys():
            self.raw_formula = re.sub(column, sheet_column_to_letters[column], self.raw_formula)

    def get_formula_with_row_number(self, row):
        formula_with_row = re.sub(r'(?<=[^A-Z])[A-Z](?=[^A-Z])',  lambda g: g.group(0) + f"{row}", self.raw_formula)

        if self.verbose:
            print(f"formula with row: {formula_with_row}")

        return formula_with_row


    